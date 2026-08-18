from __future__ import annotations

from dataclasses import dataclass, replace
from datetime import UTC, date, datetime, timedelta
from io import BytesIO
from zipfile import ZipFile

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook
from openpyxl.worksheet.table import Table

from garmin_sheets_sync.adapters.onedrive_xlsx_destination import (
    OneDriveXlsxDestination,
    RemoteFile,
)
from garmin_sheets_sync.adapters.workbook_contract import (
    ACTIVITY_HEADERS,
    DAILY_HEADERS,
    EXCEL_ACTIVITY_TABLE,
    EXCEL_DAILY_TABLE,
    EXCEL_WEIGHT_TABLE,
    WEIGHT_HEADERS,
)
from garmin_sheets_sync.errors import SchemaError
from garmin_sheets_sync.models import (
    Activity,
    DailyActivity,
    IngestionBatch,
    WeightMeasurement,
    parse_timestamp,
)

WORKBOOK_PATH = "/Apps/Garmin/RP Cut.xlsx"
COMPLETED_AT = datetime(2026, 8, 9, 20, tzinfo=UTC)


@dataclass(frozen=True, slots=True)
class Replacement:
    path: str
    content: bytes
    expected_etag: str


class FakeStorage:
    def __init__(self, content: bytes, *, etag: str = "etag-1") -> None:
        self.content = content
        self.etag = etag
        self.downloads: list[str] = []
        self.replacements: list[Replacement] = []

    def download(self, path: str) -> RemoteFile:
        self.downloads.append(path)
        return RemoteFile(content=self.content, etag=self.etag)

    def replace(self, path: str, content: bytes, expected_etag: str) -> None:
        self.replacements.append(Replacement(path, content, expected_etag))
        self.content = content
        self.etag = f"etag-{len(self.replacements) + 1}"


def _batch() -> IngestionBatch:
    return IngestionBatch(
        weights=(
            WeightMeasurement(
                measured_at=parse_timestamp("2026-08-09T06:00:00Z", "test"),
                weight_kg=82.4,
            ),
        ),
        daily_activity=(DailyActivity(date(2026, 8, 9), 1000, 123),),
        activities=(
            Activity(
                activity_id="12345678901234567",
                name="Run",
                activity_type="running",
                started_at=parse_timestamp("2026-08-09T07:00:00Z", "test"),
                duration_seconds=120,
                distance_meters=500,
                calories_kcal=42,
                average_heart_rate_bpm=135,
                max_heart_rate_bpm=162,
            ),
        ),
    )


def _workbook_bytes(workbook: OpenpyxlWorkbook | None = None) -> bytes:
    if workbook is None:
        workbook = Workbook()
        weight_log = workbook.active
        weight_log.title = "Weight Log"
        weight_log.append(WEIGHT_HEADERS)
        workbook.create_sheet("Garmin Daily Activity").append(DAILY_HEADERS)
        workbook.create_sheet("Garmin Activities").append(ACTIVITY_HEADERS)
        settings = workbook.create_sheet("Settings")
        settings["A2"] = "Last Successful Sync"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _load(content: bytes) -> OpenpyxlWorkbook:
    return load_workbook(BytesIO(content), data_only=False)


def _destination(
    workbook: OpenpyxlWorkbook | None = None,
) -> tuple[OneDriveXlsxDestination, FakeStorage]:
    storage = FakeStorage(_workbook_bytes(workbook))
    return OneDriveXlsxDestination(storage, workbook_path=WORKBOOK_PATH), storage


def _sam_diet_workbook() -> OpenpyxlWorkbook:
    workbook = Workbook()
    weight = workbook.active
    weight.title = "Weight Log"
    weight.append(["Weight Log"])
    weight.append([])
    weight.append(
        [
            "Date",
            "Timestamp",
            "Weight (lb)",
            "Weight (kg)",
            "7-Day Avg (lb)",
            "Body Fat (%)",
            "Muscle Mass (lb)",
            "Bone Mass (lb)",
            "Body Water (%)",
            "BMI",
            "Source",
            "Sync Timestamp",
            "Goal Min (lb)",
            "Goal Max (lb)",
            "Pace Slow (lb)",
            "Pace Fast (lb)",
            None,
            None,
            "Band Lower (fast edge)",
            "Band Height",
        ]
    )
    weight.append(
        [
            None,
            None,
            "=D4*Settings!$B$17",
            None,
            "=AVERAGE(C4:C4)",
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            "=Settings!$B$12",
            "=Settings!$B$13",
            "=C4",
            "=C4",
            None,
            None,
            "=P4",
            "=O4-P4",
        ]
    )
    weight["A4"].number_format = "ddd, dd mmm yyyy"
    weight["B4"].number_format = "hh:mm"
    weight["L4"].number_format = "dd mmm yyyy hh:mm"
    weight.add_table(Table(displayName=EXCEL_WEIGHT_TABLE, ref="A3:P4"))

    daily = workbook.create_sheet("Garmin Daily Activity")
    daily.append(["Garmin Daily Activity"])
    daily.append([])
    daily.append(["Date", "Steps", "Active Calories", "Total Calories", "Notes/Metadata"])
    daily.append([None, None, None, None, None])
    daily["A4"].number_format = "ddd, dd mmm yyyy"
    daily.add_table(Table(displayName=EXCEL_DAILY_TABLE, ref="A3:E4"))

    activities = workbook.create_sheet("Garmin Activities")
    activities.append(["Garmin Activities"])
    activities.append([])
    activities.append(
        [
            "Date",
            "Activity Type",
            "Activity Name",
            "Duration",
            "Distance",
            "Start Time",
            "Active Calories",
            "Garmin Activity ID",
            "Garmin Connect Link",
            "Notes/Metadata",
        ]
    )
    activities.append([None] * 10)
    activities["A4"].number_format = "ddd, dd mmm yyyy"
    activities["F4"].number_format = "hh:mm"
    activities.add_table(Table(displayName=EXCEL_ACTIVITY_TABLE, ref="A3:J4"))

    settings = workbook.create_sheet("Settings")
    settings["B12"] = 154
    settings["B13"] = 159
    settings["B17"] = 2.2046226218
    return workbook


def _header_columns(workbook: OpenpyxlWorkbook, sheet: str) -> dict[str, int]:
    worksheet = workbook[sheet]
    return {str(cell.value): cell.column for cell in worksheet[1] if cell.value is not None}


def test_onedrive_destination_uses_existing_sam_diet_tables() -> None:
    destination, storage = _destination(_sam_diet_workbook())

    first = destination.sync(_batch(), COMPLETED_AT)
    second = destination.sync(_batch(), COMPLETED_AT + timedelta(minutes=5))

    assert first.total.inserted == 3
    assert second.total.unchanged == 3
    workbook = _load(storage.content)
    weight = workbook["Weight Log"]
    assert weight["A4"].value == datetime(2026, 8, 9)
    assert weight["B4"].value == datetime(2026, 8, 9, 6)
    assert weight["C4"].value == "=D4*Settings!$B$17"
    assert weight["D4"].value == 82.4
    assert weight["K4"].value == "Garmin"
    assert weight["L4"].value == datetime(2026, 8, 9, 20)
    assert weight.tables[EXCEL_WEIGHT_TABLE].ref == "A3:P4"

    daily = workbook["Garmin Daily Activity"]
    assert daily["A4"].value == datetime(2026, 8, 9)
    assert daily["B4"].value == 1000
    assert daily["C4"].value == 123

    activities = workbook["Garmin Activities"]
    assert activities["A4"].value == datetime(2026, 8, 9)
    assert activities["D4"].value == 120
    assert activities["F4"].value == datetime(2026, 8, 9, 7)
    assert activities["H4"].value == "12345678901234567"
    assert activities["H4"].data_type == "s"
    assert workbook["Settings"]["B2"].value == "2026-08-09T20:05:00Z"


def test_sam_diet_table_growth_copies_formulas_and_filters() -> None:
    destination, storage = _destination(_sam_diet_workbook())
    initial = _batch()
    second_batch = IngestionBatch(
        weights=(
            initial.weights[0],
            replace(
                initial.weights[0],
                measured_at=parse_timestamp("2026-08-10T06:00:00Z", "test"),
                weight_kg=81.9,
            ),
        ),
        daily_activity=(
            initial.daily_activity[0],
            replace(initial.daily_activity[0], date=date(2026, 8, 10), steps=2000),
        ),
        activities=(
            initial.activities[0],
            replace(
                initial.activities[0],
                activity_id="another-activity",
                started_at=parse_timestamp("2026-08-10T07:00:00Z", "test"),
            ),
        ),
    )

    destination.sync(second_batch, COMPLETED_AT)

    workbook = _load(storage.content)
    weight = workbook["Weight Log"]
    assert weight.tables[EXCEL_WEIGHT_TABLE].ref == "A3:P5"
    assert weight.tables[EXCEL_WEIGHT_TABLE].autoFilter.ref == "A3:P5"
    assert weight["C5"].value == "=D5*Settings!$B$17"
    assert weight["E5"].value == "=AVERAGE(C5:C5)"
    assert weight["S5"].value == "=P5"
    assert weight["T5"].value == "=O5-P5"
    assert workbook["Garmin Daily Activity"].tables[EXCEL_DAILY_TABLE].ref == "A3:E5"
    assert workbook["Garmin Activities"].tables[EXCEL_ACTIVITY_TABLE].ref == "A3:J5"


def test_onedrive_destination_inserts_and_reruns_without_duplicates() -> None:
    destination, storage = _destination()

    first = destination.sync(_batch(), COMPLETED_AT)
    second = destination.sync(_batch(), COMPLETED_AT)

    assert first.total.inserted == 3
    assert second.total.unchanged == 3
    workbook = _load(storage.content)
    assert workbook["Weight Log"].max_row == 2
    assert workbook["Garmin Daily Activity"].max_row == 2
    assert workbook["Garmin Activities"].max_row == 2
    assert workbook["Settings"]["B2"].value == "2026-08-09T20:00:00Z"


def test_onedrive_destination_updates_changed_owned_values() -> None:
    destination, storage = _destination()
    initial = _batch()
    destination.sync(initial, COMPLETED_AT)
    updated = IngestionBatch(
        weights=(replace(initial.weights[0], weight_kg=81.9),),
        daily_activity=(replace(initial.daily_activity[0], steps=2000),),
        activities=(replace(initial.activities[0], name="Trail Run"),),
    )

    report = destination.sync(updated, COMPLETED_AT)

    assert report.total.updated == 3
    workbook = _load(storage.content)
    weight_columns = _header_columns(workbook, "Weight Log")
    daily_columns = _header_columns(workbook, "Garmin Daily Activity")
    activity_columns = _header_columns(workbook, "Garmin Activities")
    assert workbook["Weight Log"].cell(2, weight_columns["Weight (kg)"]).value == 81.9
    assert workbook["Garmin Daily Activity"].cell(2, daily_columns["Steps"]).value == 2000
    activity_name = workbook["Garmin Activities"].cell(2, activity_columns["Activity Name"])
    assert activity_name.value == "Trail Run"


def test_onedrive_destination_refuses_to_replace_manual_weight() -> None:
    workbook = _load(_workbook_bytes())
    workbook["Weight Log"].append(
        ["2026-08-09T06:00:00Z", 82.5, None, None, None, None, None, "manual"]
    )
    destination, storage = _destination(workbook)

    with pytest.raises(SchemaError, match="non-Garmin"):
        destination.sync(_batch(), COMPLETED_AT)

    assert storage.replacements == []


def test_onedrive_destination_requires_headers_before_uploading() -> None:
    workbook = _load(_workbook_bytes())
    worksheet = workbook["Garmin Activities"]
    missing_headers = {
        "Calories (kcal)",
        "Average Heart Rate (bpm)",
        "Max Heart Rate (bpm)",
    }
    for cell in worksheet[1]:
        if cell.value in missing_headers:
            cell.value = None
    destination, storage = _destination(workbook)

    with pytest.raises(SchemaError) as error:
        destination.sync(_batch(), COMPLETED_AT)

    assert all(header in str(error.value) for header in missing_headers)
    assert storage.replacements == []


def test_onedrive_destination_preserves_formulas_and_non_owned_cells() -> None:
    workbook = _load(_workbook_bytes())
    worksheet = workbook["Weight Log"]
    worksheet["I1"] = "Coach Notes"
    worksheet["J1"] = "Derived Weight"
    worksheet.append(
        [
            "2026-08-09T06:00:00Z",
            83,
            None,
            None,
            None,
            None,
            None,
            "Garmin",
            "Keep this note",
            "=B2*2",
        ]
    )
    destination, storage = _destination(workbook)

    destination.sync(_batch(), COMPLETED_AT)

    updated = _load(storage.content)["Weight Log"]
    assert updated["B2"].value == 82.4
    assert updated["I2"].value == "Keep this note"
    assert updated["J2"].value == "=B2*2"
    assert updated["J2"].data_type == "f"


def test_onedrive_destination_stores_long_activity_id_as_text() -> None:
    destination, storage = _destination()

    destination.sync(_batch(), COMPLETED_AT)

    workbook = _load(storage.content)
    columns = _header_columns(workbook, "Garmin Activities")
    cell = workbook["Garmin Activities"].cell(2, columns["Activity ID"])
    assert cell.value == "12345678901234567"
    assert cell.data_type == "s"


def test_onedrive_destination_rejects_numeric_existing_activity_id() -> None:
    workbook = _load(_workbook_bytes())
    workbook["Garmin Activities"].append(
        [
            12345678901234567,
            "Run",
            "running",
            "2026-08-08T07:00:00Z",
            120,
            500,
            42,
            135,
            162,
            "https://connect.garmin.com/modern/activities",
        ]
    )
    destination, storage = _destination(workbook)

    with pytest.raises(SchemaError, match="numeric Activity ID"):
        destination.sync(_batch(), COMPLETED_AT)

    assert storage.replacements == []


def test_onedrive_destination_writes_garmin_strings_as_literal_text() -> None:
    destination, storage = _destination()
    initial = _batch()
    batch = IngestionBatch(
        weights=initial.weights,
        daily_activity=initial.daily_activity,
        activities=(replace(initial.activities[0], name='=HYPERLINK("https://evil.test")'),),
    )

    destination.sync(batch, COMPLETED_AT)

    workbook = _load(storage.content)
    columns = _header_columns(workbook, "Garmin Activities")
    cell = workbook["Garmin Activities"].cell(2, columns["Activity Name"])
    assert cell.value == '=HYPERLINK("https://evil.test")'
    assert cell.data_type == "s"


def test_onedrive_destination_forwards_downloaded_etag_when_replacing() -> None:
    destination, storage = _destination()

    destination.sync(_batch(), COMPLETED_AT)
    destination.sync(_batch(), COMPLETED_AT)

    assert storage.downloads == [WORKBOOK_PATH, WORKBOOK_PATH]
    assert [replacement.path for replacement in storage.replacements] == [
        WORKBOOK_PATH,
        WORKBOOK_PATH,
    ]
    assert [replacement.expected_etag for replacement in storage.replacements] == [
        "etag-1",
        "etag-2",
    ]


def test_onedrive_destination_preserves_chart_package_parts() -> None:
    workbook = _load(_workbook_bytes())
    dashboard = workbook.create_sheet("Dashboard")
    dashboard.append(["Day", "Weight"])
    dashboard.append([1, 180])
    dashboard.append([2, 179])
    chart = BarChart()
    chart.add_data(Reference(dashboard, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    dashboard.add_chart(chart, "D2")
    destination, storage = _destination(workbook)

    destination.sync(_batch(), COMPLETED_AT)

    with ZipFile(BytesIO(storage.content)) as archive:
        names = archive.namelist()
    assert sum(name.startswith("xl/charts/") for name in names) == 1
    assert sum(name.startswith("xl/drawings/") for name in names) >= 1


def test_onedrive_destination_extends_table_and_translates_formula() -> None:
    workbook = _load(_workbook_bytes())
    worksheet = workbook["Garmin Activities"]
    worksheet.append(
        [
            "another-activity",
            "Walk",
            "walking",
            "2026-08-08T07:00:00Z",
            300,
            1000,
            30,
            100,
            120,
            "https://connect.garmin.com/modern/activities",
            "=F2/1000",
        ]
    )
    worksheet["K1"] = "Distance (km)"
    worksheet.add_table(Table(displayName="GarminActivities", ref="A1:K2"))
    destination, storage = _destination(workbook)

    destination.sync(_batch(), COMPLETED_AT)

    updated = _load(storage.content)["Garmin Activities"]
    assert updated.tables["GarminActivities"].ref == "A1:K3"
    assert updated.tables["GarminActivities"].autoFilter.ref == "A1:K3"
    assert updated["K3"].value == "=F3/1000"


def test_onedrive_destination_rejects_table_totals_rows() -> None:
    workbook = _load(_workbook_bytes())
    worksheet = workbook["Garmin Daily Activity"]
    worksheet.append(["2026-08-08", 100, 20])
    worksheet.append(["Totals", "=SUM(B2:B2)", "=SUM(C2:C2)"])
    table = Table(displayName="GarminDaily", ref="A1:C3")
    table.totalsRowShown = True
    table.totalsRowCount = 1
    worksheet.add_table(table)
    destination, storage = _destination(workbook)

    with pytest.raises(SchemaError, match="unsupported table totals row"):
        destination.sync(_batch(), COMPLETED_AT)

    assert storage.replacements == []
