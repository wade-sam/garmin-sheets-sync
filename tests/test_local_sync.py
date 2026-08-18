import json
import sqlite3
from datetime import UTC, date, datetime
from pathlib import Path

from pytest import MonkeyPatch

from garmin_sheets_sync import cli
from garmin_sheets_sync.adapters.fixture_source import FixtureSource
from garmin_sheets_sync.adapters.garmin_source import GarminConnectSource
from garmin_sheets_sync.adapters.onedrive_xlsx_destination import OneDriveXlsxDestination
from garmin_sheets_sync.adapters.sqlite_destination import SqliteDestination
from garmin_sheets_sync.cli import main
from garmin_sheets_sync.locking import RunLock
from garmin_sheets_sync.models import DateWindow
from garmin_sheets_sync.service import SyncService

FIXTURE = Path(__file__).parents[1] / "fixtures" / "sample.json"
COMPLETED_AT = datetime(2026, 8, 9, 20, 0, tzinfo=UTC)


def test_repeated_local_sync_is_idempotent(tmp_path: Path) -> None:
    database = tmp_path / "sync.db"
    service = SyncService(
        FixtureSource(FIXTURE),
        SqliteDestination(database),
        clock=lambda: COMPLETED_AT,
    )
    window = DateWindow(date(2026, 8, 8), date(2026, 8, 9))

    first = service.run(window)
    second = service.run(window)

    assert first.total.inserted == 6
    assert second.total.inserted == 0
    assert second.total.updated == 0
    assert second.total.unchanged == 6
    with sqlite3.connect(database) as connection:
        assert connection.execute("SELECT COUNT(*) FROM weight_log").fetchone() == (2,)
        assert connection.execute("SELECT COUNT(*) FROM daily_activity").fetchone() == (2,)
        assert connection.execute("SELECT COUNT(*) FROM activities").fetchone() == (2,)
        assert connection.execute(
            """
            SELECT calories_kcal, average_heart_rate_bpm, max_heart_rate_bpm
            FROM activities WHERE activity_id = '12345678901234567'
            """
        ).fetchone() == (612.0, 149.0, 176.0)
        assert connection.execute(
            """
            SELECT calories_kcal, average_heart_rate_bpm, max_heart_rate_bpm
            FROM activities WHERE activity_id = '12345678901234568'
            """
        ).fetchone() == (None, None, None)
        assert connection.execute(
            "SELECT value FROM metadata WHERE key = 'last_successful_sync'"
        ).fetchone() == ("2026-08-09T20:00:00Z",)


def test_cli_fixture_mode_needs_no_credentials(
    tmp_path: Path, monkeypatch: MonkeyPatch
) -> None:
    for name in (
        "GARMIN_EMAIL",
        "GARMIN_PASSWORD",
        "GOOGLE_SERVICE_ACCOUNT_FILE",
        "GOOGLE_SHEET_ID",
        "ONEDRIVE_CLIENT_ID",
        "ONEDRIVE_WORKBOOK_PATH",
    ):
        monkeypatch.delenv(name, raising=False)

    result = main(
        [
            "sync",
            "--source",
            "fixture",
            "--destination",
            "sqlite",
            "--fixture",
            str(FIXTURE),
            "--database",
            str(tmp_path / "local.db"),
            "--lock-file",
            str(tmp_path / "sync.lock"),
            "--start",
            "2026-08-08",
            "--end",
            "2026-08-09",
        ]
    )

    assert result == 0


def test_cli_garmin_source_uses_cached_token_without_credentials(
    tmp_path: Path, monkeypatch: MonkeyPatch
) -> None:
    token_dir = tmp_path / "garmin"
    token_dir.mkdir()
    (token_dir / "garmin_tokens.json").write_text("cached-token")
    monkeypatch.setenv("GARMIN_TOKEN_DIR", str(token_dir))
    monkeypatch.delenv("GARMIN_EMAIL", raising=False)
    monkeypatch.delenv("GARMIN_PASSWORD", raising=False)
    login_arguments: list[tuple[str, str, Path]] = []
    expected_source = object()

    def fake_login(email: str, password: str, path: Path, **_kwargs: object) -> object:
        login_arguments.append((email, password, path))
        return expected_source

    monkeypatch.setattr(
        GarminConnectSource,
        "login",
        staticmethod(fake_login),
    )
    args = cli.build_parser().parse_args(["sync", "--source", "garmin"])

    source = cli._build_source(args)

    assert source is expected_source
    assert login_arguments == [("", "", token_dir)]


def test_cli_worker_needs_no_credentials(monkeypatch: MonkeyPatch) -> None:
    for name in (
        "GARMIN_EMAIL",
        "GARMIN_PASSWORD",
        "GOOGLE_SERVICE_ACCOUNT_FILE",
        "GOOGLE_SHEET_ID",
        "ONEDRIVE_CLIENT_ID",
        "ONEDRIVE_WORKBOOK_PATH",
    ):
        monkeypatch.delenv(name, raising=False)
    worker_started = False

    def run_worker() -> int:
        nonlocal worker_started
        worker_started = True
        return 0

    monkeypatch.setattr(cli, "_run_worker", run_worker)

    result = main(["worker"])

    assert result == 0
    assert worker_started is True


def test_cli_builds_onedrive_destination_without_google_credentials(
    monkeypatch: MonkeyPatch,
) -> None:
    monkeypatch.setenv("ONEDRIVE_CLIENT_ID", "client-id")
    monkeypatch.setenv("ONEDRIVE_WORKBOOK_PATH", "/Apps/Garmin/RP Cut.xlsx")
    monkeypatch.delenv("GOOGLE_SERVICE_ACCOUNT_FILE", raising=False)
    monkeypatch.delenv("GOOGLE_SHEET_ID", raising=False)
    args = cli.build_parser().parse_args(["sync", "--destination", "onedrive"])

    destination = cli._build_destination(args)

    assert isinstance(destination, OneDriveXlsxDestination)
    assert destination.name == "onedrive"


def test_cli_onedrive_auth_runs_explicit_device_login(
    tmp_path: Path,
    monkeypatch: MonkeyPatch,
) -> None:
    authenticated = False

    class FakeProvider:
        def authenticate_device_code(self) -> None:
            nonlocal authenticated
            authenticated = True

    monkeypatch.setattr(cli, "_onedrive_token_provider", FakeProvider)

    result = main(
        [
            "onedrive-auth",
            "--lock-file",
            str(tmp_path / "sync.lock"),
        ]
    )

    assert result == 0
    assert authenticated is True


def test_cli_onedrive_inspect_is_read_only(
    tmp_path: Path,
    monkeypatch: MonkeyPatch,
    capsys: object,
) -> None:
    from io import BytesIO

    from openpyxl import Workbook

    from garmin_sheets_sync.adapters.onedrive_xlsx_destination import RemoteFile
    from garmin_sheets_sync.adapters.workbook_contract import (
        ACTIVITY_HEADERS,
        DAILY_HEADERS,
        WEIGHT_HEADERS,
    )

    workbook = Workbook()
    workbook.active.title = "Weight Log"
    workbook["Weight Log"].append(WEIGHT_HEADERS)
    workbook.create_sheet("Garmin Daily Activity").append(DAILY_HEADERS)
    workbook.create_sheet("Garmin Activities").append(ACTIVITY_HEADERS)
    workbook.create_sheet("Settings")
    output = BytesIO()
    workbook.save(output)

    class ReadOnlyStorage:
        downloads: list[str] = []

        def download(self, path: str) -> RemoteFile:
            self.downloads.append(path)
            return RemoteFile(output.getvalue(), '"etag"')

    storage = ReadOnlyStorage()
    monkeypatch.setenv("ONEDRIVE_WORKBOOK_PATH", "/Apps/Garmin/RP Cut.xlsx")
    monkeypatch.setattr(cli, "_onedrive_storage", lambda: storage)

    result = main(
        [
            "onedrive-inspect",
            "--lock-file",
            str(tmp_path / "sync.lock"),
        ]
    )

    captured = capsys.readouterr()  # type: ignore[attr-defined]
    report = json.loads(captured.out)
    assert result == 0
    assert report["ready_for_sync"] is True
    assert storage.downloads == ["/Apps/Garmin/RP Cut.xlsx"]


def test_cli_acquires_lock_before_building_source(
    tmp_path: Path, monkeypatch: MonkeyPatch
) -> None:
    lock_file = tmp_path / "sync.lock"

    source_was_built = False

    def unexpected_source_build(args: object) -> object:
        nonlocal source_was_built
        source_was_built = True
        raise AssertionError("source must not be built while another run owns the lock")

    monkeypatch.setattr(cli, "_build_source", unexpected_source_build)
    with RunLock(lock_file):
        result = main(
            [
                "sync",
                "--source",
                "fixture",
                "--destination",
                "sqlite",
                "--fixture",
                str(FIXTURE),
                "--database",
                str(tmp_path / "local.db"),
                "--lock-file",
                str(lock_file),
                "--start",
                "2026-08-08",
                "--end",
                "2026-08-09",
            ]
        )

    assert result == 1
    assert source_was_built is False


def test_sqlite_destination_migrates_legacy_activity_table(tmp_path: Path) -> None:
    database = tmp_path / "legacy.db"
    with sqlite3.connect(database) as connection:
        connection.execute(
            """
            CREATE TABLE activities (
                activity_id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                activity_type TEXT NOT NULL,
                started_at TEXT NOT NULL,
                duration_seconds REAL NOT NULL,
                distance_meters REAL,
                connect_url TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            INSERT INTO activities (
                activity_id, name, activity_type, started_at,
                duration_seconds, distance_meters, connect_url
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "legacy-activity",
                "Legacy Run",
                "running",
                "2026-08-01T06:00:00Z",
                1200,
                4000,
                "https://connect.garmin.com/modern/activities",
            ),
        )

    service = SyncService(
        FixtureSource(FIXTURE),
        SqliteDestination(database),
        clock=lambda: COMPLETED_AT,
    )
    window = DateWindow(date(2026, 8, 8), date(2026, 8, 9))

    first = service.run(window)
    second = service.run(window)

    assert first.activities.inserted == 2
    assert second.total.unchanged == 6
    with sqlite3.connect(database) as connection:
        columns = {
            row[1] for row in connection.execute("PRAGMA table_info(activities)")
        }
        assert {
            "calories_kcal",
            "average_heart_rate_bpm",
            "max_heart_rate_bpm",
        } <= columns
        assert connection.execute(
            """
            SELECT calories_kcal, average_heart_rate_bpm, max_heart_rate_bpm
            FROM activities WHERE activity_id = '12345678901234567'
            """
        ).fetchone() == (612.0, 149.0, 176.0)
        assert connection.execute(
            "SELECT name FROM activities WHERE activity_id = 'legacy-activity'"
        ).fetchone() == ("Legacy Run",)


def test_cli_rejects_live_run_with_implicit_log_only_alerting(
    tmp_path: Path, monkeypatch: MonkeyPatch
) -> None:
    monkeypatch.setenv("ALERT_MODE", "log")

    result = main(
        [
            "sync",
            "--source",
            "garmin",
            "--destination",
            "sqlite",
            "--lock-file",
            str(tmp_path / "sync.lock"),
            "--start",
            "2026-08-09",
            "--end",
            "2026-08-09",
        ]
    )

    assert result == 1
