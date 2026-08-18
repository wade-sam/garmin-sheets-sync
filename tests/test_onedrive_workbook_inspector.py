from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.worksheet.table import Table

from garmin_sheets_sync.adapters.onedrive_workbook_inspector import inspect_workbook
from garmin_sheets_sync.adapters.workbook_contract import (
    ACTIVITY_HEADERS,
    DAILY_HEADERS,
    EXCEL_ACTIVITY_TABLE,
    EXCEL_DAILY_TABLE,
    EXCEL_WEIGHT_TABLE,
    WEIGHT_HEADERS,
)


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    weight = workbook.active
    weight.title = "Weight Log"
    weight.append(WEIGHT_HEADERS + ("Derived Weight",))
    weight.append(("2026-08-09T06:00:00Z", 82.4, None, None, None, None, None, "Garmin", "=B2"))
    workbook.create_sheet("Garmin Daily Activity").append(DAILY_HEADERS)
    workbook.create_sheet("Garmin Activities").append(ACTIVITY_HEADERS)
    workbook.create_sheet("Settings")["B2"] = "2026-08-09T20:00:00Z"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_inspection_reports_structure_without_row_values() -> None:
    inspection = inspect_workbook(_workbook_bytes())

    assert inspection.ready_for_sync is True
    assert inspection.issues == ()
    assert inspection.success_marker_present is True
    assert tuple(sheet.name for sheet in inspection.workbook_sheets) == (
        "Weight Log",
        "Garmin Daily Activity",
        "Garmin Activities",
        "Settings",
    )
    weight = inspection.sheets[0]
    assert weight.data_rows == 1
    assert weight.extra_headers == ("Derived Weight",)
    assert weight.formula_columns == ("Derived Weight",)
    assert "2026-08-09" not in str(inspection.as_dict())


def test_inspection_reports_missing_contract_and_unsupported_totals() -> None:
    workbook = Workbook()
    weight = workbook.active
    weight.title = "Weight Log"
    weight.append(WEIGHT_HEADERS)
    weight.append(("Totals", None, None, None, None, None, None, None))
    table = Table(displayName="WeightLog", ref="A1:H2")
    table.totalsRowShown = True
    table.totalsRowCount = 1
    weight.add_table(table)
    output = BytesIO()
    workbook.save(output)

    inspection = inspect_workbook(output.getvalue())

    assert inspection.ready_for_sync is False
    assert any("unsupported table totals row" in issue for issue in inspection.issues)
    assert any("Garmin Daily Activity" in issue for issue in inspection.issues)
    assert any("Garmin Activities" in issue for issue in inspection.issues)
    assert any("Settings" in issue for issue in inspection.issues)


def test_inspection_recognises_sam_diet_named_tables_on_row_three() -> None:
    workbook = Workbook()
    weight = workbook.active
    weight.title = "Weight Log"
    weight.append(["Weight Log"])
    weight.append([])
    weight.append(
        [
            "Date",
            "Timestamp",
            "Weight (lb)",
            "Weight (kg)",
            "7-Day Avg (lb)",
            "Body Fat (%)",
            "Muscle Mass (lb)",
            "Bone Mass (lb)",
            "Body Water (%)",
            "BMI",
            "Source",
            "Sync Timestamp",
            "Goal Min (lb)",
            "Goal Max (lb)",
            "Pace Slow (lb)",
            "Pace Fast (lb)",
        ]
    )
    weight.append([None] * 16)
    weight.add_table(Table(displayName=EXCEL_WEIGHT_TABLE, ref="A3:P4"))

    daily = workbook.create_sheet("Garmin Daily Activity")
    daily.append(["Garmin Daily Activity"])
    daily.append([])
    daily.append(["Date", "Steps", "Active Calories", "Total Calories", "Notes/Metadata"])
    daily.append([None] * 5)
    daily.add_table(Table(displayName=EXCEL_DAILY_TABLE, ref="A3:E4"))

    activities = workbook.create_sheet("Garmin Activities")
    activities.append(["Garmin Activities"])
    activities.append([])
    activities.append(
        [
            "Date",
            "Activity Type",
            "Activity Name",
            "Duration",
            "Distance",
            "Start Time",
            "Active Calories",
            "Garmin Activity ID",
            "Garmin Connect Link",
            "Notes/Metadata",
        ]
    )
    activities.append([None] * 10)
    activities.add_table(Table(displayName=EXCEL_ACTIVITY_TABLE, ref="A3:J4"))
    workbook.create_sheet("Settings")
    output = BytesIO()
    workbook.save(output)

    inspection = inspect_workbook(output.getvalue())

    assert inspection.ready_for_sync is True
    assert inspection.issues == ()
    assert tuple(sheet.data_rows for sheet in inspection.sheets) == (0, 0, 0)
    assert inspection.sheets[0].missing_headers == ()
    assert inspection.sheets[2].missing_headers == ()
