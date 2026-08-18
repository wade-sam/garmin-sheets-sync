"""Read-only structural inspection for an existing OneDrive workbook."""

from __future__ import annotations

from collections import Counter
from dataclasses import asdict, dataclass
from io import BytesIO
from typing import Any
from zipfile import BadZipFile, ZipFile

from garmin_sheets_sync.adapters.workbook_contract import (
    ACTIVITY_HEADERS,
    DAILY_HEADERS,
    EXCEL_ACTIVITY_HEADERS,
    EXCEL_ACTIVITY_TABLE,
    EXCEL_DAILY_HEADERS,
    EXCEL_DAILY_TABLE,
    EXCEL_WEIGHT_HEADERS,
    EXCEL_WEIGHT_TABLE,
    WEIGHT_HEADERS,
)
from garmin_sheets_sync.errors import ConfigurationError, SchemaError

SHEET_CONTRACTS = (
    (
        "Weight Log",
        WEIGHT_HEADERS,
        "Measurement Timestamp",
        EXCEL_WEIGHT_HEADERS,
        "Timestamp",
        EXCEL_WEIGHT_TABLE,
    ),
    (
        "Garmin Daily Activity",
        DAILY_HEADERS,
        "Date",
        EXCEL_DAILY_HEADERS,
        "Date",
        EXCEL_DAILY_TABLE,
    ),
    (
        "Garmin Activities",
        ACTIVITY_HEADERS,
        "Activity ID",
        EXCEL_ACTIVITY_HEADERS,
        "Garmin Activity ID",
        EXCEL_ACTIVITY_TABLE,
    ),
)

PACKAGE_FEATURE_PREFIXES = {
    "charts": "xl/charts/",
    "drawings": "xl/drawings/",
    "external_links": "xl/externalLinks/",
    "media": "xl/media/",
    "pivot_tables": "xl/pivotTables/",
    "pivot_cache": "xl/pivotCache/",
    "slicers": "xl/slicers/",
}


@dataclass(frozen=True, slots=True)
class TableInspection:
    name: str
    reference: str
    auto_filter_reference: str | None
    has_totals_row: bool


@dataclass(frozen=True, slots=True)
class SheetInspection:
    name: str
    present: bool
    data_rows: int
    headers: tuple[str, ...]
    missing_headers: tuple[str, ...]
    extra_headers: tuple[str, ...]
    duplicate_headers: tuple[str, ...]
    formula_columns: tuple[str, ...]
    tables: tuple[TableInspection, ...]


@dataclass(frozen=True, slots=True)
class ExistingSheetInspection:
    name: str
    max_row: int
    max_column: int
    formula_cells: int
    tables: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class WorkbookInspection:
    ready_for_sync: bool
    issues: tuple[str, ...]
    warnings: tuple[str, ...]
    settings_tab: str
    success_marker_cell: str
    success_marker_present: bool
    workbook_sheets: tuple[ExistingSheetInspection, ...]
    sheets: tuple[SheetInspection, ...]
    package_features: dict[str, int]

    def as_dict(self) -> dict[str, Any]:
        return asdict(self)


def inspect_workbook(
    content: bytes,
    *,
    settings_tab: str = "Settings",
    last_success_cell: str = "B2",
) -> WorkbookInspection:
    """Inspect workbook structure without returning any row values."""

    try:
        with ZipFile(BytesIO(content)) as archive:
            corrupt_member = archive.testzip()
            if corrupt_member is not None:
                raise SchemaError("OneDrive workbook contains a corrupt package member")
            names = archive.namelist()
    except BadZipFile as exc:
        raise SchemaError("OneDrive workbook is not a readable .xlsx file") from exc

    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
    except ImportError as exc:
        raise ConfigurationError(
            "OneDrive workbook support is not installed; install the 'live' extra"
        ) from exc

    try:
        workbook = load_workbook(
            BytesIO(content),
            data_only=False,
            keep_links=True,
            rich_text=True,
        )
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as exc:
        raise SchemaError("OneDrive workbook is not a readable .xlsx file") from exc

    issues: list[str] = []
    warnings: list[str] = []
    sheets: list[SheetInspection] = []
    for (
        title,
        legacy_headers,
        legacy_key_header,
        table_headers,
        table_key_header,
        table_name,
    ) in SHEET_CONTRACTS:
        if title not in workbook.sheetnames:
            issues.append(f"missing required sheet {title!r}")
            sheets.append(
                SheetInspection(
                    name=title,
                    present=False,
                    data_rows=0,
                    headers=(),
                    missing_headers=table_headers,
                    extra_headers=(),
                    duplicate_headers=(),
                    formula_columns=(),
                    tables=(),
                )
            )
            continue

        worksheet = workbook[title]
        required_headers: tuple[str, ...]
        if table_name in worksheet.tables:
            from openpyxl.utils.cell import range_boundaries

            table = worksheet.tables[table_name]
            min_col, header_row, max_col, last_data_row = range_boundaries(table.ref)
            raw_headers = tuple(
                cell.value
                for cell in next(
                    worksheet.iter_rows(
                        min_row=header_row,
                        max_row=header_row,
                        min_col=min_col,
                        max_col=max_col,
                    )
                )
            )
            header_columns = {
                str(cell.value): cell.column
                for cell in next(
                    worksheet.iter_rows(
                        min_row=header_row,
                        max_row=header_row,
                        min_col=min_col,
                        max_col=max_col,
                    )
                )
                if cell.value not in (None, "")
            }
            required_headers = tuple(table_headers)
            key_header = table_key_header
            first_data_row = header_row + 1
        else:
            raw_headers = tuple(cell.value for cell in worksheet[1])
            header_columns = {
                str(value): index + 1
                for index, value in enumerate(raw_headers)
                if value not in (None, "")
            }
            required_headers = tuple(legacy_headers)
            key_header = legacy_key_header
            header_row = 1
            first_data_row = 2
            last_data_row = worksheet.max_row
        headers = tuple(
            str(value) for value in raw_headers if value not in (None, "")
        )
        counts = Counter(headers)
        duplicates = tuple(sorted(header for header, count in counts.items() if count > 1))
        missing = tuple(header for header in required_headers if header not in headers)
        extras = tuple(header for header in headers if header not in required_headers)
        for header in duplicates:
            issues.append(f"sheet {title!r} has duplicate header {header!r}")
        if missing:
            issues.append(f"sheet {title!r} is missing headers: {', '.join(missing)}")

        formula_columns = tuple(
            header
            for header, column in header_columns.items()
            if any(
                worksheet.cell(row=row, column=column).data_type == "f"
                for row in range(first_data_row, last_data_row + 1)
            )
        )
        tables = tuple(
            TableInspection(
                name=table.name,
                reference=table.ref,
                auto_filter_reference=(
                    table.autoFilter.ref if table.autoFilter is not None else None
                ),
                has_totals_row=bool(table.totalsRowShown or table.totalsRowCount),
            )
            for table in worksheet.tables.values()
        )
        if any(table.has_totals_row for table in tables):
            issues.append(f"sheet {title!r} has an unsupported table totals row")

        key_column = header_columns.get(key_header)
        data_rows = (
            sum(
                worksheet.cell(row=row, column=key_column).value not in (None, "")
                for row in range(first_data_row, last_data_row + 1)
            )
            if key_column is not None
            else 0
        )
        if key_header in {"Activity ID", "Garmin Activity ID"} and key_column is not None:
            numeric_ids = sum(
                isinstance(worksheet.cell(row=row, column=key_column).value, int | float)
                and not isinstance(worksheet.cell(row=row, column=key_column).value, bool)
                for row in range(first_data_row, last_data_row + 1)
            )
            if numeric_ids:
                issues.append(
                    f"sheet {title!r} has {numeric_ids} numeric Activity ID cell(s); "
                    "restore exact IDs and format them as text"
                )

        sheets.append(
            SheetInspection(
                name=title,
                present=True,
                data_rows=data_rows,
                headers=headers,
                missing_headers=missing,
                extra_headers=extras,
                duplicate_headers=duplicates,
                formula_columns=formula_columns,
                tables=tables,
            )
        )

    success_marker_present = False
    if settings_tab not in workbook.sheetnames:
        issues.append(f"missing required settings sheet {settings_tab!r}")
    else:
        try:
            success_marker_present = workbook[settings_tab][last_success_cell].value not in (
                None,
                "",
            )
        except ValueError as exc:
            raise SchemaError(
                f"invalid OneDrive success marker cell {last_success_cell!r}"
            ) from exc

    package_features = {
        label: sum(name.startswith(prefix) for name in names)
        for label, prefix in PACKAGE_FEATURE_PREFIXES.items()
    }
    present_features = [
        f"{label}={count}" for label, count in package_features.items() if count
    ]
    if present_features:
        warnings.append(
            "workbook contains OOXML features requiring manual post-sync verification: "
            + ", ".join(present_features)
        )

    workbook_sheets = tuple(
        ExistingSheetInspection(
            name=worksheet.title,
            max_row=worksheet.max_row,
            max_column=worksheet.max_column,
            formula_cells=sum(
                cell.data_type == "f"
                for row in worksheet.iter_rows()
                for cell in row
            ),
            tables=tuple(table.name for table in worksheet.tables.values()),
        )
        for worksheet in workbook.worksheets
    )

    workbook.close()
    return WorkbookInspection(
        ready_for_sync=not issues,
        issues=tuple(issues),
        warnings=tuple(warnings),
        settings_tab=settings_tab,
        success_marker_cell=last_success_cell,
        success_marker_present=success_marker_present,
        workbook_sheets=workbook_sheets,
        sheets=tuple(sheets),
        package_features=package_features,
    )
