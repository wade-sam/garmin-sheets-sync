"""OneDrive destination that safely updates a complete ``.xlsx`` workbook."""

from __future__ import annotations

import re
from contextlib import suppress
from copy import copy
from dataclasses import dataclass
from datetime import UTC, date, datetime
from io import BytesIO
from pathlib import PurePosixPath
from typing import Any, Protocol
from zipfile import ZIP_DEFLATED, BadZipFile, ZipFile

from garmin_sheets_sync.adapters.workbook_contract import (
    ACTIVITY_HEADERS,
    DAILY_HEADERS,
    EXCEL_ACTIVITY_HEADERS,
    EXCEL_ACTIVITY_TABLE,
    EXCEL_DAILY_HEADERS,
    EXCEL_DAILY_TABLE,
    EXCEL_WEIGHT_FORMULA_HEADERS,
    EXCEL_WEIGHT_HEADERS,
    EXCEL_WEIGHT_TABLE,
    WEIGHT_HEADERS,
)
from garmin_sheets_sync.errors import ConfigurationError, SchemaError
from garmin_sheets_sync.models import IngestionBatch, format_timestamp
from garmin_sheets_sync.ports import SyncReport, UpsertCounts

PACKAGE_FEATURE_PREFIXES = {
    "charts": "xl/charts/",
    "drawings": "xl/drawings/",
    "external links": "xl/externalLinks/",
    "media": "xl/media/",
    "pivot caches": "xl/pivotCache/",
    "pivot tables": "xl/pivotTables/",
    "slicers": "xl/slicers/",
}


@dataclass(frozen=True, slots=True)
class RemoteFile:
    content: bytes
    etag: str


class OneDriveStorage(Protocol):
    def download(self, path: str) -> RemoteFile: ...

    def replace(self, path: str, content: bytes, expected_etag: str) -> None: ...


def normalize_workbook_path(workbook_path: str) -> str:
    normalized_path = str(PurePosixPath("/" + workbook_path.strip("/")))
    if not workbook_path.strip("/") or not normalized_path.lower().endswith(".xlsx"):
        raise ConfigurationError("ONEDRIVE_WORKBOOK_PATH must identify an .xlsx file")
    return normalized_path


def _equivalent(existing: Any, expected: Any) -> bool:
    if expected is None:
        return existing in (None, "")
    if isinstance(expected, datetime) and isinstance(existing, datetime):
        return _excel_datetime(existing) == _excel_datetime(expected)
    if isinstance(expected, date) and not isinstance(expected, datetime):
        if isinstance(existing, datetime):
            existing = existing.date()
        return bool(existing == expected)
    if isinstance(expected, int | float) and not isinstance(expected, bool):
        try:
            return float(existing) == float(expected)
        except (TypeError, ValueError):
            return False
    return str(existing) == str(expected)


def _excel_datetime(value: datetime) -> datetime:
    """Return an Excel-compatible, timezone-free UTC datetime."""

    if value.tzinfo is not None:
        return value.astimezone(UTC).replace(tzinfo=None)
    return value


def _key_value(value: Any, *, date_only: bool = False) -> str:
    if isinstance(value, datetime):
        if date_only:
            return _excel_datetime(value).date().isoformat()
        return _excel_datetime(value).isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip() if value not in (None, "") else ""


def _kg_to_lb(value: float | None) -> float | None:
    return value * 2.2046226218 if value is not None else None


def _package_feature_counts(content: bytes) -> dict[str, int]:
    try:
        with ZipFile(BytesIO(content)) as archive:
            names = archive.namelist()
    except BadZipFile as exc:
        raise SchemaError("OneDrive workbook is not a readable .xlsx file") from exc
    return {
        label: sum(name.startswith(prefix) for name in names)
        for label, prefix in PACKAGE_FEATURE_PREFIXES.items()
    }


def _verify_package_features(source: bytes, rendered: bytes) -> None:
    before = _package_feature_counts(source)
    after = _package_feature_counts(rendered)
    lost = [
        f"{label} ({before[label]} -> {after[label]})"
        for label in before
        if after[label] < before[label]
    ]
    if lost:
        raise SchemaError(
            "generated OneDrive workbook lost OOXML features: " + ", ".join(lost)
        )


def _xml_element(xml: bytes, local_name: str) -> bytes:
    name = re.escape(local_name.encode())
    prefix = rb"(?:[A-Za-z_][A-Za-z0-9_.-]*:)?"
    pattern = re.compile(
        rb"<"
        + prefix
        + name
        + rb"\b[^>]*(?:/>|>.*?</"
        + prefix
        + name
        + rb"\s*>)",
        re.DOTALL,
    )
    matches = pattern.findall(xml)
    if len(matches) != 1:
        raise SchemaError(f"workbook XML has an invalid {local_name} element")
    return bytes(matches[0])


def _replace_xml_element(source: bytes, rendered: bytes, local_name: str) -> bytes:
    source_element = _xml_element(source, local_name)
    rendered_element = _xml_element(rendered, local_name)
    return source.replace(source_element, rendered_element, 1)


def _table_reference(xml: bytes) -> bytes:
    table = re.search(rb"<(?:[A-Za-z_][A-Za-z0-9_.-]*:)?table\b[^>]*>", xml)
    if table is None:
        raise SchemaError("workbook table XML has no table element")
    reference = re.search(rb'\bref="([^"]+)"', table.group())
    if reference is None:
        raise SchemaError("workbook table XML has no range reference")
    return reference.group(1)


def _replace_table_reference(source: bytes, rendered: bytes) -> bytes:
    reference = _table_reference(rendered)
    table = re.search(rb"<(?:[A-Za-z_][A-Za-z0-9_.-]*:)?table\b[^>]*>", source)
    if table is None:
        raise SchemaError("workbook table XML has no table element")
    updated_table = re.sub(
        rb'(\bref=")[^"]+(\")', rb"\g<1>" + reference + rb"\g<2>", table.group(), count=1
    )
    updated = source[: table.start()] + updated_table + source[table.end() :]

    source_filter = re.search(
        rb"<(?:[A-Za-z_][A-Za-z0-9_.-]*:)?autoFilter\b[^>]*>", updated
    )
    rendered_filter = re.search(
        rb"<(?:[A-Za-z_][A-Za-z0-9_.-]*:)?autoFilter\b[^>]*>", rendered
    )
    if source_filter is not None and rendered_filter is not None:
        filter_reference = re.search(rb'\bref="([^"]+)"', rendered_filter.group())
        if filter_reference is not None:
            updated_filter = re.sub(
                rb'(\bref=")[^"]+(\")',
                rb"\g<1>" + filter_reference.group(1) + rb"\g<2>",
                source_filter.group(),
                count=1,
            )
            updated = (
                updated[: source_filter.start()]
                + updated_filter
                + updated[source_filter.end() :]
            )
    return updated


def _style_count(archive: ZipFile) -> int:
    styles = archive.read("xl/styles.xml")
    match = re.search(rb"<(?:[A-Za-z_][A-Za-z0-9_.-]*:)?cellXfs\b[^>]*\bcount=\"(\d+)\"", styles)
    if match is None:
        raise SchemaError("workbook styles XML has no cell style count")
    return int(match.group(1))


def _merge_rendered_package(
    source: bytes,
    rendered: bytes,
    *,
    sheet_parts: frozenset[str],
    table_parts: frozenset[str],
) -> bytes:
    """Merge only managed cells/table ranges into the original OOXML package."""

    output = BytesIO()
    try:
        with (
            ZipFile(BytesIO(source)) as source_archive,
            ZipFile(BytesIO(rendered)) as rendered_archive,
            ZipFile(output, mode="w", compression=ZIP_DEFLATED) as merged_archive,
        ):
            style_count = _style_count(source_archive)
            for part in sheet_parts | table_parts:
                if part not in source_archive.namelist() or part not in rendered_archive.namelist():
                    raise SchemaError(f"workbook package is missing managed part {part!r}")
            for info in source_archive.infolist():
                content = source_archive.read(info.filename)
                if info.filename in sheet_parts:
                    rendered_sheet = rendered_archive.read(info.filename)
                    sheet_data = _xml_element(rendered_sheet, "sheetData")
                    style_ids = [int(value) for value in re.findall(rb'\bs="(\d+)"', sheet_data)]
                    if style_ids and max(style_ids) >= style_count:
                        raise SchemaError(
                            "generated OneDrive workbook requires a new cell style; "
                            "format the template data rows before syncing"
                        )
                    content = _replace_xml_element(content, rendered_sheet, "sheetData")
                    with suppress(SchemaError):
                        content = _replace_xml_element(content, rendered_sheet, "dimension")
                elif info.filename in table_parts:
                    content = _replace_table_reference(
                        content, rendered_archive.read(info.filename)
                    )
                merged_archive.writestr(info, content)
            merged_archive.comment = source_archive.comment
    except (BadZipFile, KeyError, OSError, ValueError) as exc:
        raise SchemaError("could not merge the generated OneDrive workbook") from exc
    return output.getvalue()


class OneDriveXlsxDestination:
    """Download, modify, and conditionally replace a personal OneDrive workbook."""

    name = "onedrive"

    def __init__(
        self,
        storage: OneDriveStorage,
        *,
        workbook_path: str,
        settings_tab: str = "Settings",
        last_success_cell: str = "B2",
    ) -> None:
        self._storage = storage
        self._workbook_path = normalize_workbook_path(workbook_path)
        self._settings_tab = settings_tab
        self._last_success_cell = last_success_cell

    def sync(self, batch: IngestionBatch, completed_at: datetime) -> SyncReport:
        try:
            from openpyxl import load_workbook
            from openpyxl.utils.exceptions import InvalidFileException
        except ImportError as exc:
            raise ConfigurationError(
                "OneDrive workbook support is not installed; install the 'live' extra"
            ) from exc

        remote = self._storage.download(self._workbook_path)
        try:
            workbook = load_workbook(
                BytesIO(remote.content),
                data_only=False,
                keep_links=True,
                rich_text=True,
            )
        except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as exc:
            raise SchemaError("OneDrive workbook is not a readable .xlsx file") from exc

        if self._uses_excel_template(workbook):
            sync_timestamp = _excel_datetime(completed_at)
            weights = self._upsert_sheet(
                workbook,
                "Weight Log",
                EXCEL_WEIGHT_HEADERS,
                "Timestamp",
                [
                    {
                        "Date": _excel_datetime(item.measured_at).date(),
                        "Timestamp": _excel_datetime(item.measured_at),
                        "Weight (kg)": item.weight_kg,
                        "Body Fat (%)": item.body_fat_percent,
                        "Muscle Mass (lb)": _kg_to_lb(item.skeletal_muscle_mass_kg),
                        "Bone Mass (lb)": _kg_to_lb(item.bone_mass_kg),
                        "Body Water (%)": item.body_water_percent,
                        "BMI": item.bmi,
                        "Source": item.source,
                        "Sync Timestamp": sync_timestamp,
                    }
                    for item in batch.weights
                ],
                table_name=EXCEL_WEIGHT_TABLE,
                protect_manual_source=True,
                formula_headers=EXCEL_WEIGHT_FORMULA_HEADERS,
                volatile_headers=frozenset({"Sync Timestamp"}),
            )
            daily = self._upsert_sheet(
                workbook,
                "Garmin Daily Activity",
                EXCEL_DAILY_HEADERS,
                "Date",
                [
                    {
                        "Date": item.date,
                        "Steps": item.steps,
                        "Active Calories": item.active_calories,
                    }
                    for item in batch.daily_activity
                ],
                table_name=EXCEL_DAILY_TABLE,
            )
            activities = self._upsert_sheet(
                workbook,
                "Garmin Activities",
                EXCEL_ACTIVITY_HEADERS,
                "Garmin Activity ID",
                [
                    {
                        "Date": _excel_datetime(item.started_at).date(),
                        "Activity Type": item.activity_type,
                        "Activity Name": item.name,
                        "Duration": item.duration_seconds,
                        "Distance": item.distance_meters,
                        "Start Time": _excel_datetime(item.started_at),
                        "Active Calories": item.calories_kcal,
                        "Garmin Activity ID": item.key,
                        "Garmin Connect Link": item.connect_url,
                    }
                    for item in batch.activities
                ],
                table_name=EXCEL_ACTIVITY_TABLE,
                text_headers=frozenset({"Garmin Activity ID"}),
            )
        else:
            weights = self._upsert_sheet(
                workbook,
                "Weight Log",
                WEIGHT_HEADERS,
                "Measurement Timestamp",
                [
                    {
                        "Measurement Timestamp": item.key,
                        "Weight (kg)": item.weight_kg,
                        "Body Fat (%)": item.body_fat_percent,
                        "Skeletal Muscle Mass (kg)": item.skeletal_muscle_mass_kg,
                        "Bone Mass (kg)": item.bone_mass_kg,
                        "Body Water (%)": item.body_water_percent,
                        "BMI": item.bmi,
                        "Source": item.source,
                    }
                    for item in batch.weights
                ],
                protect_manual_source=True,
            )
            daily = self._upsert_sheet(
                workbook,
                "Garmin Daily Activity",
                DAILY_HEADERS,
                "Date",
                [
                    {
                        "Date": item.key,
                        "Steps": item.steps,
                        "Active Calories": item.active_calories,
                    }
                    for item in batch.daily_activity
                ],
            )
            activities = self._upsert_sheet(
                workbook,
                "Garmin Activities",
                ACTIVITY_HEADERS,
                "Activity ID",
                [
                    {
                        "Activity ID": item.key,
                        "Activity Name": item.name,
                        "Activity Type": item.activity_type,
                        "Start Time": format_timestamp(item.started_at),
                        "Duration (seconds)": item.duration_seconds,
                        "Distance (meters)": item.distance_meters,
                        "Calories (kcal)": item.calories_kcal,
                        "Average Heart Rate (bpm)": item.average_heart_rate_bpm,
                        "Max Heart Rate (bpm)": item.max_heart_rate_bpm,
                        "Garmin Connect Link": item.connect_url,
                    }
                    for item in batch.activities
                ],
                text_headers=frozenset({"Activity ID"}),
            )

        settings = self._worksheet(workbook, self._settings_tab)
        try:
            settings[self._last_success_cell] = format_timestamp(completed_at)
        except ValueError as exc:
            raise SchemaError(
                f"invalid OneDrive success marker cell {self._last_success_cell!r}"
            ) from exc

        calculation = getattr(workbook, "calculation", None)
        if calculation is not None:
            calculation.calcMode = "auto"
            calculation.fullCalcOnLoad = True
            calculation.forceFullCalc = True

        managed_worksheets = tuple(
            self._worksheet(workbook, title)
            for title in (
                "Weight Log",
                "Garmin Daily Activity",
                "Garmin Activities",
                self._settings_tab,
            )
        )
        output = BytesIO()
        workbook.save(output)
        sheet_parts = frozenset(
            worksheet.path.lstrip("/") for worksheet in managed_worksheets
        )
        table_parts = frozenset(
            table.path.lstrip("/")
            for worksheet in managed_worksheets
            for table in worksheet.tables.values()
        )
        rendered_workbook = _merge_rendered_package(
            remote.content,
            output.getvalue(),
            sheet_parts=sheet_parts,
            table_parts=table_parts,
        )
        _verify_package_features(remote.content, rendered_workbook)
        try:
            verification = load_workbook(
                BytesIO(rendered_workbook),
                read_only=True,
                data_only=False,
                keep_links=True,
            )
            verification.close()
        except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as exc:
            raise SchemaError("generated OneDrive workbook could not be reopened") from exc

        self._storage.replace(self._workbook_path, rendered_workbook, remote.etag)
        return SyncReport(
            weights=weights,
            daily_activity=daily,
            activities=activities,
            completed_at=completed_at,
        )

    @staticmethod
    def _worksheet(workbook: Any, title: str) -> Any:
        if title not in workbook.sheetnames:
            raise SchemaError(f"required OneDrive workbook sheet {title!r} is unavailable")
        return workbook[title]

    @classmethod
    def _uses_excel_template(cls, workbook: Any) -> bool:
        try:
            from openpyxl.utils.cell import range_boundaries
        except ImportError:
            return False

        expected_tables = (
            ("Weight Log", EXCEL_WEIGHT_TABLE),
            ("Garmin Daily Activity", EXCEL_DAILY_TABLE),
            ("Garmin Activities", EXCEL_ACTIVITY_TABLE),
        )
        present = []
        for title, table_name in expected_tables:
            if title not in workbook.sheetnames or table_name not in workbook[title].tables:
                present.append(False)
                continue
            _, header_row, _, _ = range_boundaries(
                workbook[title].tables[table_name].ref
            )
            present.append(header_row == 3)
        if any(present) and not all(present):
            missing = [
                table_name
                for (_, table_name), is_present in zip(
                    expected_tables, present, strict=True
                )
                if not is_present
            ]
            raise SchemaError(
                "OneDrive workbook has an incomplete Excel table contract; missing: "
                + ", ".join(missing)
            )
        return all(present)

    @classmethod
    def _upsert_sheet(
        cls,
        workbook: Any,
        title: str,
        required_headers: tuple[str, ...],
        key_header: str,
        records: list[dict[str, Any]],
        *,
        table_name: str | None = None,
        protect_manual_source: bool = False,
        text_headers: frozenset[str] = frozenset(),
        formula_headers: tuple[str, ...] = (),
        volatile_headers: frozenset[str] = frozenset(),
    ) -> UpsertCounts:
        worksheet = cls._worksheet(workbook, title)
        table = None
        if table_name is not None:
            try:
                from openpyxl.utils.cell import range_boundaries

                table = worksheet.tables[table_name]
                min_col, header_row_number, max_col, last_data_row = range_boundaries(
                    table.ref
                )
            except KeyError as exc:
                raise SchemaError(
                    f"OneDrive workbook sheet {title!r} is missing table {table_name!r}"
                ) from exc
            header_cells = next(
                worksheet.iter_rows(
                    min_row=header_row_number,
                    max_row=header_row_number,
                    min_col=min_col,
                    max_col=max_col,
                )
            )
            first_data_row = header_row_number + 1
        else:
            header_row_number = 1
            header_cells = next(worksheet.iter_rows(min_row=1, max_row=1))
            first_data_row = 2
            last_data_row = worksheet.max_row
        header_row = [cell.value for cell in header_cells]
        populated_headers = [str(header) for header in header_row if header not in (None, "")]
        if not populated_headers:
            raise SchemaError(f"OneDrive workbook sheet {title!r} has no header row")
        if len(populated_headers) != len(set(populated_headers)):
            raise SchemaError(f"OneDrive workbook sheet {title!r} has duplicate headers")
        header_columns = {
            str(cell.value): cell.column
            for cell in header_cells
            if cell.value not in (None, "")
        }
        missing = [header for header in required_headers if header not in header_columns]
        if missing:
            raise SchemaError(
                f"OneDrive workbook sheet {title!r} is missing headers: {', '.join(missing)}"
            )
        checked_tables = (table,) if table is not None else tuple(worksheet.tables.values())
        for checked_table in checked_tables:
            if checked_table.totalsRowShown or checked_table.totalsRowCount:
                raise SchemaError(
                    f"OneDrive workbook sheet {title!r} has an unsupported table totals row"
                )

        key_column = header_columns[key_header]
        existing_by_key: dict[str, int] = {}
        last_key_row = header_row_number
        for row_number in range(first_data_row, last_data_row + 1):
            raw_key = worksheet.cell(row=row_number, column=key_column).value
            if (
                key_header in text_headers
                and isinstance(raw_key, int | float)
                and not isinstance(raw_key, bool)
            ):
                raise SchemaError(
                    f"OneDrive workbook sheet {title!r} contains a numeric {key_header}; "
                    "restore the exact identifier and format it as text"
                )
            key = _key_value(raw_key, date_only=key_header == "Date")
            if not key:
                continue
            if key in existing_by_key:
                raise SchemaError(
                    f"OneDrive workbook sheet {title!r} contains duplicate key {key!r}"
                )
            existing_by_key[key] = row_number
            last_key_row = row_number

        counts = UpsertCounts()
        next_row = max(first_data_row, last_key_row + 1)
        for record in records:
            key = _key_value(record[key_header], date_only=key_header == "Date")
            existing_row = existing_by_key.get(key)
            if existing_row is None:
                row_number = next_row
                next_row += 1
                owned_columns = {header_columns[header] for header in required_headers}
                if table is not None:
                    cls._extend_named_table_for_row(
                        worksheet, table, row_number, owned_columns
                    )
                    cls._copy_formula_headers(
                        worksheet,
                        header_row_number,
                        row_number,
                        formula_headers,
                    )
                else:
                    cls._extend_tables_for_row(worksheet, row_number, owned_columns)
                counts = counts + UpsertCounts(inserted=1)
            else:
                row_number = existing_row
                if protect_manual_source:
                    existing_source = worksheet.cell(
                        row=row_number, column=header_columns["Source"]
                    ).value
                    if existing_source != "Garmin":
                        raise SchemaError(
                            f"refusing to overwrite non-Garmin Weight Log row for {key!r}"
                        )
                changed = any(
                    not _equivalent(
                        worksheet.cell(row=row_number, column=header_columns[header]).value,
                        record[header],
                    )
                    for header in required_headers
                    if header not in volatile_headers
                )
                counts = counts + (
                    UpsertCounts(updated=1) if changed else UpsertCounts(unchanged=1)
                )
                if not changed:
                    continue

            for header in required_headers:
                cell = worksheet.cell(row=row_number, column=header_columns[header])
                cell.value = record[header]
                if isinstance(record[header], str):
                    cell.data_type = "s"
            existing_by_key[key] = row_number
        return counts

    @staticmethod
    def _extend_named_table_for_row(
        worksheet: Any, table: Any, row_number: int, owned_columns: set[int]
    ) -> None:
        try:
            from openpyxl.formula.translate import Translator
            from openpyxl.utils.cell import get_column_letter, range_boundaries
        except ImportError:
            return

        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        if row_number <= max_row:
            return
        if row_number != max_row + 1:
            raise SchemaError(
                f"cannot extend OneDrive workbook table {table.name!r} across blank rows"
            )
        for column in range(min_col, max_col + 1):
            source = worksheet.cell(row=max_row, column=column)
            target = worksheet.cell(row=row_number, column=column)
            if source.has_style:
                target._style = copy(source._style)
            target.number_format = source.number_format
            if column not in owned_columns and source.data_type == "f":
                target.value = Translator(
                    source.value,
                    origin=source.coordinate,
                ).translate_formula(target.coordinate)
        table.ref = (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{row_number}"
        )
        if table.autoFilter is not None:
            table.autoFilter.ref = table.ref

    @staticmethod
    def _copy_formula_headers(
        worksheet: Any,
        header_row_number: int,
        row_number: int,
        formula_headers: tuple[str, ...],
    ) -> None:
        if not formula_headers or row_number <= header_row_number + 1:
            return
        try:
            from openpyxl.formula.translate import Translator
        except ImportError:
            return

        formula_header_set = set(formula_headers)
        columns = {
            str(cell.value): cell.column
            for cell in worksheet[header_row_number]
            if cell.value in formula_header_set
        }
        for header in formula_headers:
            column = columns.get(header)
            if column is None:
                continue
            source = worksheet.cell(row=row_number - 1, column=column)
            target = worksheet.cell(row=row_number, column=column)
            if source.has_style:
                target._style = copy(source._style)
            target.number_format = source.number_format
            if source.data_type == "f":
                target.value = Translator(
                    source.value,
                    origin=source.coordinate,
                ).translate_formula(target.coordinate)

    @staticmethod
    def _extend_tables_for_row(
        worksheet: Any, row_number: int, owned_columns: set[int]
    ) -> None:
        try:
            from openpyxl.formula.translate import Translator
            from openpyxl.utils.cell import get_column_letter, range_boundaries
        except ImportError:
            return

        for table in worksheet.tables.values():
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            if min_row != 1 or not owned_columns.issubset(range(min_col, max_col + 1)):
                continue
            if row_number <= max_row:
                continue
            for column in range(min_col, max_col + 1):
                if column in owned_columns or max_row < 2:
                    continue
                source = worksheet.cell(row=max_row, column=column)
                target = worksheet.cell(row=row_number, column=column)
                if isinstance(source.value, str) and source.value.startswith("="):
                    target.value = Translator(
                        source.value,
                        origin=source.coordinate,
                    ).translate_formula(target.coordinate)
                    if source.has_style:
                        target._style = copy(source._style)
                    target.number_format = source.number_format
            table.ref = (
                f"{get_column_letter(min_col)}{min_row}:"
                f"{get_column_letter(max_col)}{row_number}"
            )
            if table.autoFilter is not None:
                table.autoFilter.ref = table.ref
